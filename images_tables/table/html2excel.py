# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
html_content = """
<table style="min-width: 800px;"><colgroup><col style="min-width: 100px;"><col style="min-width: 100px;"><col style="min-width: 100px;"><col style="min-width: 100px;"><col style="min-width: 100px;"><col style="min-width: 100px;"><col style="min-width: 100px;"><col style="min-width: 100px;"></colgroup><tbody><tr><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>绕组</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>线饼</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>辐向压力 (N/mm)</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>压曲强度 (N/mm)</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>安全系数</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>拉伸应力 (Mpa)</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>许用应力 (Mpa)</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>安全系数</p></td></tr><tr><td colspan="1" rowspan="3" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>中压</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>A</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>-94.91</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>739.61</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>7.79</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td></tr><tr><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>B</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>-80.45</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>493.08</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>6.13</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td></tr><tr><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>C</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>-87.98</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>493.08</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>5.60</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td></tr><tr><td colspan="1" rowspan="5" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>高压</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>A</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>45.49</p></td><td colspan="1" rowspan="5" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td><td colspan="1" rowspan="5" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p></p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>40.94</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>160.0</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>3.91</p></td></tr><tr><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>B</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>55.19</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>49.67</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>160.0</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>3.22</p></td></tr><tr><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>E</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>86.66</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>66.75</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>160.0</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>2.40</p></td></tr><tr><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>F</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>57.48</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>44.27</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>160.0</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>3.61</p></td></tr><tr><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>G</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>49.50</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>69.38</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>160.0</p></td><td colspan="1" rowspan="1" data-border-top="true" data-border-right="true" data-border-bottom="true" data-border-left="true" text-align="left" style="text-align: left;"><p>2.31</p></td></tr></tbody></table>
"""


def html_to_excel_openpyxl(html_str, output_filename):
    # 1. 解析 HTML
    soup = BeautifulSoup(html_str, 'html.parser')
    table = soup.find('table')

    wb = Workbook()
    ws = wb.active
    ws.title = "table"

    # 用于跟踪单元格占用情况的集合 (row, col)
    occupied = set()

    for r_idx, tr in enumerate(table.find_all('tr'), start=1):
        c_idx = 1
        for td in tr.find_all(['td', 'th']):
            # 跳过已被 rowspan 占用的位置
            while (r_idx, c_idx) in occupied:
                c_idx += 1

            # 获取跨度属性
            rowspan = int(td.get('rowspan', 1))
            colspan = int(td.get('colspan', 1))
            value = td.get_text(strip=True)

            # 写入值到当前起始单元格
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # 基础样式：居中和边框
            thin = Side(border_style="thin", color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # 处理合并单元格
            if rowspan > 1 or colspan > 1:
                end_row = r_idx + rowspan - 1
                end_col = c_idx + colspan - 1
                ws.merge_cells(start_row=r_idx, start_column=c_idx,
                               end_row=end_row, end_column=end_col)

                # 标记被占用的格子
                for r in range(r_idx, end_row + 1):
                    for c in range(c_idx, end_col + 1):
                        occupied.add((r, c))
            else:
                occupied.add((r_idx, c_idx))

            c_idx += colspan

    # 自动调整列宽
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        # 找到该列中所有单元格内容的最大长度
        # 注意：合并单元格的值只存在于其左上角的第一个格子中，其他格点为 None
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    # 简单计算中文字符长度（中文占位比英文宽，这里 *2 处理）
                    val_str = str(cell.value)
                    byte_len = len(val_str.encode('gbk'))
                    if byte_len > max_length:
                        max_length = byte_len
            except:
                pass

        # 使用 get_column_letter 转换索引数字为字母（如 1 -> 'A'）
        column_letter = get_column_letter(col_idx)
        # 设置宽度，最小不低于 10
        ws.column_dimensions[column_letter].width = max(max_length + 2, 10)

    wb.save(output_filename)
    print(f"Excel文件已生成: {output_filename}")

if __name__ == "__main__":
    # 调用函数
    html_to_excel_openpyxl(html_content, "线路参数明细3.xlsx")